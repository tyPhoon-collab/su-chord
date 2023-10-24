"""
解析用のExcelファイルのデータ部分を置換するためのスクリプト
CLIとして扱う想定

ex) python3 python/excel_data_replacer.py 'test/outputs/result.xlsx'
"""

import argparse
import csv
import glob
import itertools
import os

import natsort
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

DEFAULT_OUTPUT_PATH = "test/outputs/result.xlsx"
START_COLUMN = 2
START_ROW = 3


class ExcelDataReplacer:
    def __init__(self, paths: list[str], start_row: int = 1, start_column: int = 1) -> None:
        self.paths = paths
        self.start_row = start_row
        self.start_column = start_column

    @staticmethod
    def _read_csv(path: str) -> list[list[str]]:
        with open(path) as f:
            data = [row for row in csv.reader(f)]

        return data

    @staticmethod
    def _get_workbook(output_path: str) -> Workbook:
        # wb = load_workbook("test/outputs/result3.xlsx")
        if os.path.exists(output_path):
            wb = load_workbook(output_path)
        else:
            wb = Workbook()
        return wb

    def _write(self, ws: Worksheet, path: str, offset_row: int) -> None:
        data_rows = self._read_csv(path)

        # fmt: off
        rows = ws.iter_rows(
            min_row=offset_row + self.start_row,
            max_row=offset_row + self.start_row + 5 * 13 + 1,
            min_col=self.start_column,
            max_col=self.start_column + 21,
        )
        # fmt: on
        for row, data_row in zip(rows, data_rows):
            for i in range(len(data_row)):
                cell = row[i]
                value = data_row[i]
                cell.value = value

    def write(self, output_path: str) -> None:
        assert output_path.endswith(".xlsx")

        wb = self._get_workbook(output_path)
        ws: Worksheet = wb["Sheet1"]

        for i in range(len(self.paths)):
            path = self.paths[i]
            # print(f"creating {path}...")

            self._write(ws, path, offset_row=i * (5 * 13 + 2))

        wb.save(output_path)
        print(f"create {output_path} done!")


def _get_files(input_path: str) -> list[str]:
    """
    ファイルパスとディレクトリパスのどちらも対応する
    """
    if input_path.endswith(".csv"):
        print("detect file path")
        return [input_path]

    print("detect directory path")
    files = glob.glob(f"{input_path}/*.csv")
    return natsort.natsorted(files)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="csv to excel")

    parser.add_argument("input_path", nargs="+", help="input path. file or dir")
    parser.add_argument("-o", "--output_path", help="output path", default=DEFAULT_OUTPUT_PATH)

    args = parser.parse_args()

    paths_list = [_get_files(path) for path in args.input_path]
    paths = list(itertools.chain.from_iterable(paths_list))

    ExcelDataReplacer(paths, start_column=START_COLUMN, start_row=START_ROW).write(args.output_path)
